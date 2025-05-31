
import json 
import csv
data = {
    "Date":["2024-05-01","2024-05-02","2024-05-03"],
    "Product":["Laptop","Mobile","Chair"],
    "Category":["Electronics","Electronics","Furniture"],
    "Region":["North","South","East"],
    "Units Sold":[20,30,15],
    "Unit Price":[50000,20000,1500]

}
with open("data.json","w")as file:
    json.dump(data,file,indent=4)

print("data write successfully in data.json file")
try:
    with open("data.json","r")as file:
        content = json.load(file)
except(FileNotFoundError,json.JSONDecodeError):
    print("file not found error")
    content = []

content = [dict(zip(content.keys(),t))for t in zip(*content.values())]
with open("Sales.csv","w",newline="")as file:
    writter = csv.DictWriter(file,fieldnames=content[0].keys())
    writter.writeheader()
    writter.writerows(content)
print("data load csv file successfully")
import pandas as pd 
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
df = pd.read_csv("Sales.csv")
print(df)
df["Total Revenue"] = df["Units Sold"] * df["Unit Price"]
with pd.ExcelWriter("ProductSales.xlsx",engine="openpyxl")as writter:
    df.to_excel(writter,sheet_name="Sheet1",index=False)
    df.head().to_excel(writter,sheet_name="Top5",index=False)

print("data write in excel")
wb = load_workbook("ProductSales.xlsx")
ws = wb.active
for cell in ws[1]:
    cell.font = Font(bold=True)

for coll in ws.columns:
    coll_num = coll[0].column
    coll_letter = get_column_letter(coll_num)
    max_length = 0
    for cell in coll:
        try:
            if cell.value:
                max_length = max(max_length,len(str(cell.value)))
                adjust_length = max_length+2
                ws.column_dimensions[coll_letter].width = adjust_length
        except:
            pass

wb.save("ProductSales.xlsx")